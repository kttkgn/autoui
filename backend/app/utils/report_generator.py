import os
import json
from datetime import datetime
from typing import Dict, Any, List
from jinja2 import Environment, FileSystemLoader
import pdfkit
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from app.core.config import settings
from app.models.project import TestExecution

class ReportGenerator:
    def __init__(self):
        self.template_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "templates")
        self.env = Environment(loader=FileSystemLoader(self.template_dir))
        self.report_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "reports")
        os.makedirs(self.report_dir, exist_ok=True)

    async def generate_report(self, execution: TestExecution) -> str:
        """生成单个测试执行的详细报告"""
        template = self.env.get_template("test_report.html")
        
        # 准备报告数据
        report_data = {
            "execution": execution,
            "test_case": execution.test_case,
            "test_suite": execution.test_case.test_suite,
            "project": execution.test_case.test_suite.project,
            "generated_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        # 渲染报告
        html_content = template.render(**report_data)
        
        # 保存报告
        report_path = os.path.join(self.report_dir, f"execution_{execution.id}.html")
        with open(report_path, "w", encoding="utf-8") as f:
            f.write(html_content)
        
        return f"/reports/execution_{execution.id}.html"

    async def generate_summary_report(self, project_id: int, executions: List[TestExecution]) -> str:
        """生成测试执行汇总报告"""
        template = self.env.get_template("test_summary_report.html")
        
        # 计算统计信息
        total = len(executions)
        success = sum(1 for e in executions if e.status == "success")
        failed = sum(1 for e in executions if e.status == "failed")
        running = sum(1 for e in executions if e.status == "running")
        
        # 计算平均执行时间
        completed_executions = [e for e in executions if e.duration is not None]
        avg_duration = sum(e.duration for e in completed_executions) / len(completed_executions) if completed_executions else 0
        
        # 计算成功率
        success_rate = (success / total * 100) if total > 0 else 0
        
        # 按测试套件统计
        suite_stats = {}
        for execution in executions:
            suite_name = execution.test_case.test_suite.name
            if suite_name not in suite_stats:
                suite_stats[suite_name] = {
                    "total": 0,
                    "success": 0,
                    "failed": 0,
                    "running": 0
                }
            suite_stats[suite_name]["total"] += 1
            suite_stats[suite_name][execution.status] += 1
        
        # 按日期统计
        date_stats = {}
        for execution in executions:
            date = execution.start_time.strftime("%Y-%m-%d")
            if date not in date_stats:
                date_stats[date] = {
                    "total": 0,
                    "success": 0,
                    "failed": 0,
                    "running": 0
                }
            date_stats[date]["total"] += 1
            date_stats[date][execution.status] += 1
        
        # 准备报告数据
        report_data = {
            "total": total,
            "success": success,
            "failed": failed,
            "running": running,
            "avg_duration": round(avg_duration, 2),
            "success_rate": round(success_rate, 2),
            "suite_stats": suite_stats,
            "date_stats": date_stats,
            "executions": executions,
            "generated_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        # 渲染报告
        html_content = template.render(**report_data)
        
        # 保存报告
        report_path = os.path.join(self.report_dir, f"project_{project_id}_summary.html")
        with open(report_path, "w", encoding="utf-8") as f:
            f.write(html_content)
        
        return f"/reports/project_{project_id}_summary.html"

    async def export_to_pdf(self, report_path: str) -> str:
        """将HTML报告导出为PDF"""
        try:
            # 配置wkhtmltopdf选项
            options = {
                'page-size': 'A4',
                'margin-top': '20mm',
                'margin-right': '20mm',
                'margin-bottom': '20mm',
                'margin-left': '20mm',
                'encoding': 'UTF-8',
                'no-outline': None
            }
            
            # 生成PDF文件路径
            pdf_path = report_path.replace('.html', '.pdf')
            
            # 转换HTML为PDF
            pdfkit.from_file(report_path, pdf_path, options=options)
            
            return pdf_path
        except Exception as e:
            raise Exception(f"导出PDF失败: {str(e)}")

    async def export_to_excel(self, project_id: int, executions: List[TestExecution]) -> str:
        """将测试执行数据导出为Excel"""
        try:
            # 创建工作簿和工作表
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "测试执行汇总"
            
            # 设置样式
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            center_aligned = Alignment(horizontal='center', vertical='center')
            
            # 写入表头
            headers = [
                "执行ID", "测试用例", "测试套件", "状态", "设备", 
                "开始时间", "结束时间", "执行时长(秒)", "错误信息"
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_aligned
            
            # 写入数据
            for row, execution in enumerate(executions, 2):
                ws.cell(row=row, column=1).value = execution.id
                ws.cell(row=row, column=2).value = execution.test_case.name
                ws.cell(row=row, column=3).value = execution.test_case.test_suite.name
                ws.cell(row=row, column=4).value = execution.status
                ws.cell(row=row, column=5).value = execution.device_name
                ws.cell(row=row, column=6).value = execution.start_time.strftime("%Y-%m-%d %H:%M:%S")
                ws.cell(row=row, column=7).value = execution.end_time.strftime("%Y-%m-%d %H:%M:%S") if execution.end_time else ""
                ws.cell(row=row, column=8).value = execution.duration
                ws.cell(row=row, column=9).value = execution.error_message
            
            # 调整列宽
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            # 保存文件
            excel_path = os.path.join(self.report_dir, f"project_{project_id}_summary.xlsx")
            wb.save(excel_path)
            
            return excel_path
        except Exception as e:
            raise Exception(f"导出Excel失败: {str(e)}") 